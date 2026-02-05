import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BRAND_COLORS = {
    "primary": "989898",    # Cinza Escuro
    "secondary": "f1f1f1",  # Cinza Claro
    "white": "FFFFFF",      # Branco
    "text": "000000"        # Preto
}

def _get_standard_border():
    side = Side(style='thin', color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def generate_excel_report(df: pd.DataFrame, title: str, sheet_name: str = "Relatório"):
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Panda converte para Excel e começa na coluna 1
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
        
        ws = writer.sheets[sheet_name]
        num_cols = len(df.columns)

        # 2. Titulo da planilha
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1, value=title.upper())
        title_cell.font = Font(size=14, bold=True, color=BRAND_COLORS["white"])
        title_cell.fill = PatternFill(start_color=BRAND_COLORS["primary"], fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # 3. Formatação das colunas + aplicando nas colunas
        header_style = {
            "font": Font(bold=True, color=BRAND_COLORS["text"]),
            "fill": PatternFill(start_color=BRAND_COLORS["secondary"], fill_type="solid"),
            "alignment": Alignment(horizontal='center', vertical='center'),
            "border": _get_standard_border()
        }
        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=2, column=col_num)
            for key, value in header_style.items():
                setattr(cell, key, value)

        # 4. Formatação do resto das celulas
        body_alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.alignment = body_alignment
                cell.border = _get_standard_border()

        # 5. Ajuste na dimensao das colunas
        for i, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 22

    return output.getvalue()